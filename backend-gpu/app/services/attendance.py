from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import re

# folder to store all attendance files
ATTENDANCE_FOLDER = "attendance"
os.makedirs(ATTENDANCE_FOLDER, exist_ok=True)

# memory cache to prevent duplicate entries per day
marked_today = set()

# ----------------------------
# HELPER TO CLEAN FILE NAME
# ----------------------------
def sanitize_filename(name):
    """Remove forbidden characters from filenames (Windows-safe)."""
    return re.sub(r'[\\/*?:"<>|]', "_", name)

def get_file_path(class_name: str):
    safe_name = sanitize_filename(class_name)
    return os.path.join(ATTENDANCE_FOLDER, f"{safe_name}.xlsx")

# ----------------------------
# MARK ATTENDANCE
# ----------------------------
def mark_attendance(student_id: int, student_name: str, class_name: str):
    """
    Marks attendance for a student in a class.
    Prevents duplicate entries per day.
    """
    today = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")

    # unique key to prevent duplicates
    key = f"{student_id}_{class_name}_{today}"

    if key in marked_today:
        return  # already marked today

    marked_today.add(key)

    file_path = get_file_path(class_name)

    # create file if it doesn't exist
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Student ID", "Name", "Date", "Time"])
        wb.save(file_path)

    # append attendance
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([student_id, student_name, today, time_now])
    wb.save(file_path)

    print(f"✅ {student_name} (ID:{student_id}) marked in {class_name} at {time_now}")
    
